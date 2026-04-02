#!/usr/bin/env python3
"""
Eval runner — reads test cases from evals/eval_test_cases.xlsx, runs them
against the agent's test endpoint, judges the responses, and writes results.

Usage:
    python evals/run_evals.py --tag "v1.0" --desc "Initial baseline run"
    python evals/run_evals.py --tag "v1.1_intent" --desc "Tune intent prompt" --sheets "Intent Classifier"
    python evals/run_evals.py --tag "v1.0_calibration" --desc "Calibration baseline" --calibrate

Requirements:
    - Agent must be running with APP_ENV=test
    - ANTHROPIC_API_KEY must be set (for LLM judge calls)
    - openpyxl, requests, litellm installed
"""
import argparse
import asyncio
import json
import os
import sys
import time
from datetime import datetime

import requests

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import litellm

from evals.config import (  # noqa: E402
    AGENT_MODEL,
    AGENT_TEST_ENDPOINT,
    AGENT_TIMEOUT,
    AVG_AGENT_COMPLETION_TOKENS,
    AVG_AGENT_PROMPT_TOKENS,
    AVG_JUDGE_COMPLETION_TOKENS,
    AVG_JUDGE_PROMPT_TOKENS,
    DEFAULT_CUSTOMER_ID,
    EVAL_RUNS_DIR,
    EVALS_DIR,
    JUDGE_MODEL_BEHAVIORAL,
    JUDGE_MODEL_CALIBRATION,
    JUDGE_MODEL_CLASSIFICATION,
    MODEL_PRICE_PER_TOKEN,
    RUN_HISTORY_SHEET,
    SHEET_CALL_PROFILE,
    SHEET_JUDGE_TYPE,
    SHEET_NAMES,
    TEST_CASES_FILE,
)

# Lazy import openpyxl so the error message is clear if it's missing
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    sys.exit("openpyxl is required. Run: pip install openpyxl")

# Judge modules (lazy import to avoid loading LiteLLM until needed)
from evals.judges.classification import (  # noqa: E402
    judge_input_guard,
    judge_intent_classifier,
    judge_output_guard,
)
from evals.judges.behavioral import (  # noqa: E402
    judge_kb_retrieval,
    judge_action_execution,
    judge_escalation,
    judge_conversation_quality,
)
from evals.judges.safety import (  # noqa: E402
    judge_pii_leakage,
    judge_policy_compliance,
    judge_graceful_failure,
    judge_context_retention,
)


# ---------------------------------------------------------------------------
# Excel colour fills
# ---------------------------------------------------------------------------

FILL_PASS    = PatternFill("solid", fgColor="E6F4EA")  # green
FILL_PARTIAL = PatternFill("solid", fgColor="FFF3E0")  # orange
FILL_FAIL    = PatternFill("solid", fgColor="FCE4EC")  # red
FILL_HEADER  = PatternFill("solid", fgColor="E8EAF6")  # indigo-ish header
FONT_BOLD    = Font(bold=True)


# ---------------------------------------------------------------------------
# Excel layout helpers — work around openpyxl's max_row / max_column returning
# the *formatted* extent of the sheet (which includes pre-formatted empty rows /
# columns) rather than the last row / column that actually contains data.
# ---------------------------------------------------------------------------

def _true_last_row(ws) -> int:
    """Return the last row index that has at least one non-None cell, or 0."""
    for row in range(ws.max_row, 0, -1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row, col).value is not None:
                return row
    return 0


def _true_last_col(ws, header_row: int = 2) -> int:
    """Return the last column index with a non-None value in header_row, or 0."""
    for col in range(ws.max_column, 0, -1):
        if ws.cell(header_row, col).value is not None:
            return col
    return 0


def _ws_append(ws, data: list):
    """Write data as a new row immediately after the last non-empty row."""
    next_row = _true_last_row(ws) + 1
    for col, value in enumerate(data, start=1):
        ws.cell(next_row, col, value)


# ---------------------------------------------------------------------------
# Helpers — parse conversation field
# ---------------------------------------------------------------------------

_ROLE_NORMALISE = {"user": "customer", "human": "customer", "assistant": "agent", "bot": "agent"}


def _parse_conversation(raw) -> list:
    """
    The 'conversation' column can be:
      - A plain string (single customer message)
      - A JSON string encoding a list of {"role": ..., "content": ...} dicts

    Returns a list of {"role": ..., "content": ...} dicts always.
    Role values are normalised to the agent's expected format:
      user/human → customer
      assistant/bot → agent
    """
    if raw is None:
        return []
    if isinstance(raw, list):
        msgs = raw
    else:
        raw_str = str(raw).strip()
        if raw_str.startswith("["):
            try:
                msgs = json.loads(raw_str)
            except json.JSONDecodeError:
                msgs = None
        else:
            msgs = None
        if msgs is None:
            return [{"role": "customer", "content": raw_str}]

    return [
        {"role": _ROLE_NORMALISE.get(m.get("role", ""), m.get("role", "customer")),
         "content": m.get("content", "")}
        for m in msgs
    ]


def _parse_json_field(raw, default=None):
    """Parse a JSON string field; return default on failure."""
    if raw is None:
        return default if default is not None else {}
    if isinstance(raw, (dict, list)):
        return raw
    try:
        return json.loads(str(raw))
    except (json.JSONDecodeError, TypeError):
        return default if default is not None else {}


def _row_to_dict(ws, row_idx: int) -> dict:
    """Read a worksheet row into a dict keyed by the header row (row 2).

    Row 1 is the merged group-label row added for visual grouping.
    Row 2 is column headers. Data starts at row 3.
    """
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    return {
        h: ws.cell(row_idx, c).value
        for c, h in enumerate(headers, 1)
        if h is not None
    }


# ---------------------------------------------------------------------------
# KB reference content lookup
# ---------------------------------------------------------------------------

async def _fetch_kb_reference_content(titles: list) -> "str | None":
    """
    Fetch and concatenate KB chunk texts for the given article titles.
    Uses a deterministic title lookup (not similarity search).
    Returns None when titles is empty or no chunks are found.
    """
    if not titles:
        return None
    from backend.db.session import AsyncSessionLocal
    from backend.db.models import KBDocument, KBChunk
    from sqlalchemy import select

    parts = []
    async with AsyncSessionLocal() as session:
        for title in titles:
            result = await session.execute(
                select(KBChunk.chunk_text)
                .join(KBDocument, KBChunk.document_id == KBDocument.id)
                .where(KBDocument.title == title)
                .order_by(KBChunk.chunk_index)
            )
            rows = result.scalars().all()
            if rows:
                parts.append(f"=== {title} ===\n" + "\n\n".join(rows))

    return "\n\n".join(parts) if parts else None


# ---------------------------------------------------------------------------
# Agent call helpers
# ---------------------------------------------------------------------------

def _call_agent_full(messages: list, mock_context: dict, customer_id: str = None,
                     test_id: str = "", version_tag: str = "",
                     mock_agent_state: dict = None) -> dict:
    """POST to /api/chat/test and return parsed JSON response."""
    payload = {
        "customer_id": customer_id or DEFAULT_CUSTOMER_ID,
        "messages": messages,
        "mock_context": mock_context,
        "test_id": test_id,
        "version_tag": version_tag,
    }
    if mock_agent_state:
        payload["mock_agent_state"] = mock_agent_state
    try:
        resp = requests.post(AGENT_TEST_ENDPOINT, json=payload, timeout=AGENT_TIMEOUT)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.Timeout:
        return {"error": "api_timeout"}
    except requests.exceptions.ConnectionError:
        return {"error": "Agent not reachable. Is it running with APP_ENV=test?"}
    except Exception as e:
        return {"error": str(e)}


def _call_agent_output_guard(agent_response: str, tools_called: list, known_ids: dict,
                              test_id: str = "", version_tag: str = "") -> dict:
    """POST to /api/chat/test in output guard mode."""
    payload = {
        "test_output_guard": True,
        "agent_response": agent_response,
        "tools_called": tools_called,
        "known_ids": known_ids,
        "test_id": test_id,
        "version_tag": version_tag,
    }
    try:
        resp = requests.post(AGENT_TEST_ENDPOINT, json=payload, timeout=AGENT_TIMEOUT)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.Timeout:
        return {"error": "api_timeout"}
    except requests.exceptions.ConnectionError:
        return {"error": "Agent not reachable."}
    except Exception as e:
        return {"error": str(e)}


# ---------------------------------------------------------------------------
# Per-sheet run logic
# ---------------------------------------------------------------------------

async def run_input_guard(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    """Returns (agent_response, judgment)."""
    messages = [{"role": "customer", "content": test_case.get("customer_message", "")}]
    agent_resp = _call_agent_full(messages, {}, test_id=test_id, version_tag=version_tag)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    # Summarise the guard decision as the display "response" for this sheet
    if agent_resp.get("input_guard_blocked"):
        agent_resp["response"] = f"blocked: {agent_resp.get('input_guard_reason', 'unknown')}"
    else:
        agent_resp["response"] = "passed: safe"
    judgment = judge_input_guard(test_case, agent_resp)
    return agent_resp, judgment


async def run_intent_classifier(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    messages = _parse_conversation(test_case.get("conversation"))
    agent_resp = _call_agent_full(messages, {}, test_id=test_id, version_tag=version_tag)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    # Show the classified intent as the display "response" for this sheet
    agent_resp["response"] = agent_resp.get("inferred_intent", "unknown")
    judgment = judge_intent_classifier(test_case, agent_resp)
    return agent_resp, judgment


async def run_output_guard(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    agent_response_text = str(test_case.get("agent_response", ""))
    tools_called = _parse_json_field(test_case.get("tools_called"), default=[])
    known_ids = _parse_json_field(test_case.get("known_ids"), default={})

    agent_resp = _call_agent_output_guard(agent_response_text, tools_called, known_ids,
                                          test_id=test_id, version_tag=version_tag)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    # Summarise the guard decision as the display "response" for this sheet
    verdict = agent_resp.get("output_guard_verdict", "pass")
    failure_type = agent_resp.get("output_guard_failure_type", "none")
    agent_resp["response"] = f"{verdict}: {failure_type}"
    judgment = await judge_output_guard(test_case, agent_resp)
    return agent_resp, judgment


async def run_kb_retrieval(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    messages = _parse_conversation(test_case.get("conversation"))
    agent_resp = _call_agent_full(messages, {}, test_id=test_id, version_tag=version_tag)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    # Fetch actual KB article content by title and inject it for the judge.
    # reference_articles is a JSON array of article titles; [] means no relevant article exists.
    reference_titles = _parse_json_field(test_case.get("reference_articles"), default=[])
    reference_content = await _fetch_kb_reference_content(reference_titles)
    test_case_with_ref = {**test_case, "reference_content": reference_content}
    judgment = await judge_kb_retrieval(test_case_with_ref, agent_resp, calibrate)
    return agent_resp, judgment


async def run_action_execution(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    messages = _parse_conversation(test_case.get("conversation"))
    mock_context = _parse_json_field(test_case.get("mock_account_state"))
    mock_agent_state = _parse_json_field(test_case.get("mock_agent_state"))
    agent_resp = _call_agent_full(messages, mock_context, test_id=test_id, version_tag=version_tag,
                                  mock_agent_state=mock_agent_state)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    judgment = await judge_action_execution(test_case, agent_resp, calibrate)
    return agent_resp, judgment


async def run_escalation(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    messages = _parse_conversation(test_case.get("conversation"))
    mock_context = _parse_json_field(test_case.get("mock_account_state"))
    agent_resp = _call_agent_full(messages, mock_context, test_id=test_id, version_tag=version_tag)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    judgment = await judge_escalation(test_case, agent_resp, calibrate)
    return agent_resp, judgment


async def run_conversation_quality(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    """
    Conversation Quality cases use {{AGENT_RESPONSE}} as a placeholder.
    We send the preceding turns to the agent, capture the real response,
    then pass that response to the judge.
    """
    raw_conv = _parse_conversation(test_case.get("conversation"))
    # Split on the {{AGENT_RESPONSE}} placeholder
    pre_turns = []
    for msg in raw_conv:
        if "{{AGENT_RESPONSE}}" in str(msg.get("content", "")):
            break
        pre_turns.append(msg)

    messages = pre_turns
    if not messages:
        messages = [{"role": "customer", "content": "Hello"}]

    agent_resp = _call_agent_full(messages, {}, test_id=test_id, version_tag=version_tag)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    judgment = await judge_conversation_quality(test_case, agent_resp, calibrate)
    return agent_resp, judgment


async def run_pii_leakage(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    messages = _parse_conversation(test_case.get("conversation"))
    mock_context = _parse_json_field(test_case.get("mock_account_state"))
    agent_resp = _call_agent_full(messages, mock_context, test_id=test_id, version_tag=version_tag)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    judgment = await judge_pii_leakage(test_case, agent_resp, calibrate)
    return agent_resp, judgment


async def run_policy_compliance(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    messages = _parse_conversation(test_case.get("conversation"))
    mock_context = _parse_json_field(test_case.get("mock_account_state"))
    mock_agent_state = _parse_json_field(test_case.get("mock_agent_state"))
    agent_resp = _call_agent_full(messages, mock_context, test_id=test_id, version_tag=version_tag,
                                  mock_agent_state=mock_agent_state)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    judgment = await judge_policy_compliance(test_case, agent_resp, calibrate)
    return agent_resp, judgment


async def run_graceful_failure(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    # Note: simulated_failure is metadata for the judge — the test endpoint
    # doesn't inject failures at the tool level. The judge evaluates honesty
    # about failures that occur naturally (e.g. order not found in DB).
    messages = _parse_conversation(test_case.get("conversation"))
    mock_agent_state = _parse_json_field(test_case.get("mock_agent_state"))
    agent_resp = _call_agent_full(messages, {}, test_id=test_id, version_tag=version_tag,
                                  mock_agent_state=mock_agent_state)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    judgment = await judge_graceful_failure(test_case, agent_resp, calibrate)
    return agent_resp, judgment


async def run_context_retention(test_case: dict, calibrate: bool, test_id: str = "", version_tag: str = "") -> tuple[dict, dict]:
    messages = _parse_conversation(test_case.get("conversation"))
    mock_context = _parse_json_field(test_case.get("mock_account_state"))
    mock_agent_state = _parse_json_field(test_case.get("mock_agent_state"))
    # For multi-turn conversations, pass the full history; only the last message
    # is the one the agent needs to respond to.
    agent_resp = _call_agent_full(messages, mock_context, test_id=test_id, version_tag=version_tag,
                                  mock_agent_state=mock_agent_state)
    if "error" in agent_resp:
        return agent_resp, {"verdict": "fail", "score": 0.0, "reasoning": agent_resp["error"]}
    judgment = await judge_context_retention(test_case, agent_resp, calibrate)
    return agent_resp, judgment


# Extra columns appended after the standard 4 for specific sheets.
# Each entry is a list of (header_suffix, agent_resp_key) tuples.
_SHEET_EXTRA_COLS = {
    "Escalation": [("escalation_summary", "context_summary")],
}

# Map sheet name → runner function
_SHEET_RUNNERS = {
    "Input Guard":          run_input_guard,
    "Intent Classifier":    run_intent_classifier,
    "Output Guard":         run_output_guard,
    "KB Retrieval":         run_kb_retrieval,
    "Action Execution":     run_action_execution,
    "Escalation":           run_escalation,
    "Conversation Quality": run_conversation_quality,
    "PII & Data Leakage":   run_pii_leakage,
    "Policy Compliance":    run_policy_compliance,
    "Graceful Failure":     run_graceful_failure,
    "Context Retention":    run_context_retention,
}

# Sheets whose judges are programmatic (no LLM call) — skipped in --judge-only mode
PROGRAMMATIC_SHEETS = {"Input Guard", "Intent Classifier", "Output Guard"}

# Map sheet name → judge function (for --judge-only mode, which bypasses runner logic)
_JUDGE_DISPATCH = {
    "KB Retrieval":         judge_kb_retrieval,
    "Action Execution":     judge_action_execution,
    "Escalation":           judge_escalation,
    "Conversation Quality": judge_conversation_quality,
    "PII & Data Leakage":   judge_pii_leakage,
    "Policy Compliance":    judge_policy_compliance,
    "Graceful Failure":     judge_graceful_failure,
    "Context Retention":    judge_context_retention,
}


# ---------------------------------------------------------------------------
# Results writing
# ---------------------------------------------------------------------------

def _append_run_column(ws, tag: str, row_results: list, sheet_cost: float,
                       extra_cols: list = None):
    """
    Append columns to a test sheet for this run.

    Standard 6 columns:
      Row 1: merged group label (version tag)
      Row 2: "{tag} ($X.XXX)", "{tag} response", "{tag} reasoning", "{tag} failure_reason",
             "{tag} latency_s", "{tag} cost_usd"
      Row 3+: PASS/FAIL (color-coded), response text, judge reasoning, failure_reason,
              latency in seconds, per-case cost in USD

    extra_cols: optional list of (header_suffix, agent_resp_key) tuples appended after
      the standard 6. E.g. [("escalation_summary", "context_summary")] adds a 7th column
      whose values come from agent_response["context_summary"].

    Sheet layout: row 1 = group labels, row 2 = headers, row 3+ = data.
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    extra_cols = extra_cols or []
    n_cols = 6 + len(extra_cols)

    col = _true_last_col(ws) + 1
    fill_map  = {"pass": FILL_PASS, "fail": FILL_FAIL}
    label_map = {"pass": "PASS", "fail": "FAIL"}

    # Row 1: merged group label across all columns
    ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(col + n_cols - 1)}1")
    label_cell = ws.cell(1, col, tag)
    label_cell.fill = FILL_HEADER
    label_cell.font = FONT_BOLD
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: column headers
    headers = [
        f"{tag} (${sheet_cost:.3f})",
        f"{tag} response",
        f"{tag} reasoning",
        f"{tag} failure_reason",
        f"{tag} latency_s",
        f"{tag} cost_usd",
    ] + [f"{tag} {suffix}" for suffix, _ in extra_cols]
    for offset, title in enumerate(headers):
        cell = ws.cell(2, col + offset, title)
        cell.fill = FILL_HEADER
        cell.font = FONT_BOLD

    FILL_SKIP = PatternFill("solid", fgColor="F5F5F5")  # light grey for skipped rows

    for i, case in enumerate(row_results):
        row = i + 3  # data starts at row 3

        if case.get("skipped"):
            verdict_cell = ws.cell(row, col, "SKIP")
            verdict_cell.fill = FILL_SKIP
            continue

        result     = case.get("result", {})
        agent_resp = case.get("agent_response", {})
        verdict    = result.get("verdict", "fail")

        verdict_cell = ws.cell(row, col, label_map.get(verdict, "FAIL"))
        verdict_cell.fill = fill_map.get(verdict, FILL_FAIL)

        response_text = str(agent_resp.get("response", "") or "")
        ws.cell(row, col + 1, response_text[:500])

        ws.cell(row, col + 2, result.get("reasoning", ""))

        failure_reason = result.get("failure_reason") if verdict == "fail" else None
        ws.cell(row, col + 3, failure_reason or "")

        ws.cell(row, col + 4, round(case.get("latency_s", 0.0), 2))
        ws.cell(row, col + 5, round(case.get("cost_usd", 0.0), 5))

        for j, (_, key) in enumerate(extra_cols):
            ws.cell(row, col + 6 + j, str(agent_resp.get(key) or ""))


# ---------------------------------------------------------------------------
# Analysis sheet
# ---------------------------------------------------------------------------

def _build_analysis_sheet(wb):
    """
    Build (or rebuild) the Analysis sheet as the first sheet in the workbook.

    Table 1  — pass rate per eval sheet × version_tag (COUNTIF/COUNTA formulas).
    Tables 2-12 — failure reason count + % per eval sheet × version_tag.

    All values are Excel formulas so they update when test data changes.
    The sheet is recreated from scratch on every run so it always reflects
    the current set of version_tags and failure reasons.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.formatting.rule import CellIsRule

    ROW_LIMIT = 300   # upper bound for COUNTIF ranges — covers 215+ cases

    # ── 1. Collect version tags from Run History (chronological, deduplicated) ──
    tags = []
    if RUN_HISTORY_SHEET in wb.sheetnames:
        rh = wb[RUN_HISTORY_SHEET]
        vt_col = next(
            (c for c in range(1, rh.max_column + 1) if rh.cell(1, c).value == "version_tag"),
            None,
        )
        if vt_col:
            seen: set = set()
            for row in range(2, rh.max_row + 1):
                tag = rh.cell(row, vt_col).value
                if tag and tag not in seen:
                    seen.add(tag)
                    tags.append(str(tag))

    # ── 2. For each eval sheet, map tag → (verdict_col_letter, fr_col_letter) ──
    # Verdict column headers look like  "v1.0_baseline ($0.041)"
    # Failure_reason column is always verdict_col + 3.
    # Also detect the skip column letter (header == "skip") if present.
    sheet_cols: dict = {}        # {sheet_name: {tag: (verdict_letter, fr_letter)}}
    sheet_skip_col: dict = {}    # {sheet_name: skip_col_letter | None}
    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_cols[sheet_name] = {}
        sheet_skip_col[sheet_name] = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(2, c).value
            if v == "skip":
                sheet_skip_col[sheet_name] = get_column_letter(c)
                break
        for tag in tags:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(2, c).value
                if v and str(v).startswith(f"{tag} ($"):
                    sheet_cols[sheet_name][tag] = (
                        get_column_letter(c),
                        get_column_letter(c + 3),
                    )
                    break

    # ── 3. Collect distinct failure reasons per sheet (across all runs) ──
    sheet_reasons: dict = {}  # {sheet_name: [sorted reason strings]}
    for sheet_name, tag_map in sheet_cols.items():
        ws = wb[sheet_name]
        reasons: set = set()
        for _, fr_col in tag_map.values():
            fr_idx = column_index_from_string(fr_col)
            for row in range(3, ws.max_row + 1):
                v = ws.cell(row, fr_idx).value
                if v and str(v).strip():
                    reasons.add(str(v).strip())
        sheet_reasons[sheet_name] = sorted(reasons)

    # ── 4. Create / replace Analysis sheet at position 0 ──────────────────────
    if "Analysis" in wb.sheetnames:
        del wb["Analysis"]
    ws_a = wb.create_sheet("Analysis", 0)

    # Local style objects (don't collide with module-level fills)
    HDR_FILL   = PatternFill("solid", fgColor="E8EAF6")   # same as FILL_HEADER
    SEC_FILL   = PatternFill("solid", fgColor="EDE7F6")   # slightly darker for section labels
    CF_GREEN   = PatternFill("solid", fgColor="E6F4EA")   # matches FILL_PASS
    CF_AMBER   = PatternFill("solid", fgColor="FFF3E0")   # matches FILL_PARTIAL
    CF_RED     = PatternFill("solid", fgColor="FCE4EC")   # matches FILL_FAIL
    BOLD       = Font(bold=True)
    WRAP_TOP   = Alignment(wrap_text=True,  vertical="top")
    CENTER_MID = Alignment(horizontal="center", vertical="center", wrap_text=True)

    n_tags     = len(tags)
    last_col   = max(2, 1 + n_tags)   # rightmost column used

    # ── 5. Table 1 — Pass rate by eval sheet ──────────────────────────────────

    # Row 1: column headers
    h = ws_a.cell(1, 1, "Eval sheet")
    h.font = BOLD; h.fill = HDR_FILL; h.alignment = CENTER_MID

    for j, tag in enumerate(tags, start=2):
        c = ws_a.cell(1, j, tag)
        c.font = BOLD; c.fill = HDR_FILL; c.alignment = CENTER_MID

    # Rows 2-12: one per eval sheet
    for i, sheet_name in enumerate(SHEET_NAMES, start=2):
        ws_a.cell(i, 1, sheet_name).font = BOLD

        tag_map = sheet_cols.get(sheet_name, {})
        for j, tag in enumerate(tags, start=2):
            if tag not in tag_map:
                continue
            verdict_col, _ = tag_map[tag]
            sn = sheet_name.replace("'", "''")   # escape single quotes in sheet names
            skip_col = sheet_skip_col.get(sheet_name)
            if skip_col:
                formula = (
                    f"=IFERROR("
                    f"COUNTIF('{sn}'!{verdict_col}3:{verdict_col}{ROW_LIMIT},\"PASS\")"
                    f"/(COUNTA('{sn}'!A3:A{ROW_LIMIT})"
                    f"-COUNTIF('{sn}'!{skip_col}3:{skip_col}{ROW_LIMIT},\"TRUE\"))"
                    f",0)"
                )
            else:
                formula = (
                    f"=IFERROR("
                    f"COUNTIF('{sn}'!{verdict_col}3:{verdict_col}{ROW_LIMIT},\"PASS\")"
                    f"/COUNTA('{sn}'!A3:A{ROW_LIMIT})"
                    f",0)"
                )
            cell = ws_a.cell(i, j, formula)
            cell.number_format = "0%"
            cell.alignment = CENTER_MID

    # Conditional formatting on the pass-rate data block (B2:last_col × 12)
    if n_tags > 0:
        cf_range = f"B2:{get_column_letter(last_col)}{1 + len(SHEET_NAMES)}"
        ws_a.conditional_formatting.add(
            cf_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=CF_GREEN),
        )
        ws_a.conditional_formatting.add(
            cf_range,
            CellIsRule(operator="between", formula=["0.7", "0.8999"], fill=CF_AMBER),
        )
        ws_a.conditional_formatting.add(
            cf_range,
            CellIsRule(operator="lessThan", formula=["0.7"], fill=CF_RED),
        )

    # ── 6. Tables 2-12 — Failure reason breakdowns ────────────────────────────

    current_row = 1 + len(SHEET_NAMES) + 2   # blank row after Table 1

    for sheet_name in SHEET_NAMES:
        reasons = sheet_reasons.get(sheet_name, [])
        tag_map   = sheet_cols.get(sheet_name, {})
        sn        = sheet_name.replace("'", "''")
        # Only include tags that were actually run on this sheet
        sheet_tags = [t for t in tags if t in tag_map]
        n_sheet_tags = len(sheet_tags)
        table_last_col = max(1, n_sheet_tags + 1)

        # Section label (merged across columns used by this sheet's table)
        if table_last_col > 1:
            ws_a.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=table_last_col,
            )
        lbl = ws_a.cell(current_row, 1, f"Failure reasons: {sheet_name}")
        lbl.font = BOLD; lbl.fill = SEC_FILL; lbl.alignment = CENTER_MID
        current_row += 1

        # Sub-header row — only columns for tags run on this sheet
        ws_a.cell(current_row, 1, "failure_reason").font = BOLD
        ws_a.cell(current_row, 1).fill = HDR_FILL
        for j, tag in enumerate(sheet_tags, start=2):
            c = ws_a.cell(current_row, j, tag)
            c.font = BOLD; c.fill = HDR_FILL; c.alignment = CENTER_MID
        current_row += 1

        if not reasons:
            ws_a.cell(current_row, 1, "(no failure reasons recorded)")
            current_row += 2   # data row + blank separator
            continue

        for reason in reasons:
            ws_a.cell(current_row, 1, reason).alignment = WRAP_TOP
            for j, tag in enumerate(sheet_tags, start=2):
                _, fr_col = tag_map[tag]
                cnt = f"COUNTIF('{sn}'!{fr_col}3:{fr_col}{ROW_LIMIT},\"{reason}\")"
                tot = f"COUNTA('{sn}'!A3:A{ROW_LIMIT})"
                formula = f'=IFERROR({cnt}&" ("&TEXT({cnt}/{tot},"0%")&")","0 (0%)")'
                ws_a.cell(current_row, j, formula).alignment = CENTER_MID
            current_row += 1

        current_row += 1   # blank row between tables

    # ── 7. Column widths + zoom ────────────────────────────────────────────────
    ws_a.column_dimensions["A"].width = 32
    for j in range(2, last_col + 1):
        ws_a.column_dimensions[get_column_letter(j)].width = 18

    ws_a.sheet_view.zoomScale = 125


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

# Width of the extra (non-standard) columns appended by _SHEET_EXTRA_COLS
_EXTRA_COL_WIDTH = 50

# Run History column widths keyed by header string
_RH_COL_WIDTHS = {
    "run_id": 12, "date": 18, "version_tag": 16, "change_description": 45,
    "eval_type": 16, "pass%": 10, "total_tokens": 14, "total_cost_usd": 16,
    "judge_model": 20, "notes": 45,
}

# Repeating 4-column pattern for result groups on test sheets
_RESULT_COL_WIDTHS = [18, 55, 50, 30]  # score, response, reasoning, failure_reason


def _format_test_sheet(ws):
    """
    Apply aesthetic formatting to a test case sheet (all sheets except Run History).

    - Freeze panes at (row 3, first result column) so header rows and static
      test-case columns stay fixed while scrolling.
    - Result columns: wrap_text=True, vertical='top', widths follow the
      repeating 4-column pattern [score=18, response=55, reasoning=50,
      failure_reason=30]. Extra (5th+) columns per group use _EXTRA_COL_WIDTH.
    - Zoom set to 125%.

    The first result column is detected by scanning row-2 headers for the
    pattern " ($" which the runner always writes into verdict-column headers
    (e.g. "v1.0_baseline ($0.041)").
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    # Detect first result column
    first_result_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v and " ($" in str(v):
            first_result_col = c
            break

    if first_result_col is None:
        # No run columns yet — just set zoom
        ws.sheet_view.zoomScale = 125
        return

    # Freeze panes: rows 1-2 + all static columns stay fixed
    ws.freeze_panes = ws.cell(3, first_result_col)

    # Width + alignment for every result column
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for c in range(first_result_col, ws.max_column + 1):
        offset = c - first_result_col
        # Each run group = 4 standard columns + any extra columns for this sheet.
        # Within a group: positions 0-3 use _RESULT_COL_WIDTHS; position 4+
        # (extra columns like escalation_summary) use _EXTRA_COL_WIDTH.
        n_extra = len(_SHEET_EXTRA_COLS.get(ws.title, []))
        group_width = 4 + n_extra
        pos_in_group = offset % group_width
        width = _RESULT_COL_WIDTHS[pos_in_group] if pos_in_group < 4 else _EXTRA_COL_WIDTH
        ws.column_dimensions[get_column_letter(c)].width = width

        for row in range(3, ws.max_row + 1):
            ws.cell(row, c).alignment = wrap_top

    ws.sheet_view.zoomScale = 125


def _format_run_history_sheet(ws):
    """
    Apply aesthetic formatting to the Run History sheet.

    - Column widths from _RH_COL_WIDTHS.
    - judge_model column: wrap_text=False (clipped — only readable when selected).
    - All other columns: wrap_text=True, vertical='top'.
    - Zoom set to 125%.
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    # Find judge_model column index
    judge_model_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == "judge_model":
            judge_model_col = c
            break

    # Column widths (keyed by row-1 header)
    for c in range(1, ws.max_column + 1):
        header = ws.cell(1, c).value
        if header in _RH_COL_WIDTHS:
            ws.column_dimensions[get_column_letter(c)].width = _RH_COL_WIDTHS[header]

    # Cell alignment for data rows
    wrap_top  = Alignment(wrap_text=True,  vertical="top")
    clip_top  = Alignment(wrap_text=False, vertical="top")
    for row in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row, c).alignment = clip_top if c == judge_model_col else wrap_top

    ws.sheet_view.zoomScale = 125


def _ensure_run_history_sheet(wb) -> openpyxl.worksheet.worksheet.Worksheet:
    """Create or return the Run History sheet."""
    if RUN_HISTORY_SHEET in wb.sheetnames:
        return wb[RUN_HISTORY_SHEET]
    ws = wb.create_sheet(RUN_HISTORY_SHEET)
    headers = ["run_id", "date", "version_tag", "change_description",
               "eval_type", "pass%", "total_tokens", "total_cost_usd", "judge_model", "notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_BOLD
    return ws


def _append_run_history(ws, run_id: int, tag: str, desc: str,
                         sheet_pass_rates: dict, sheet_costs: dict,
                         sheet_tokens: dict, calibrate: bool):
    """
    Append one row per evaluated sheet plus an OVERALL row.
    Columns: run_id | date | version_tag | change_description | eval_type | pass% |
             total_tokens | total_cost_usd | judge_model | notes
    """
    judge_model = (
        JUDGE_MODEL_CALIBRATION if calibrate
        else f"classification: {JUDGE_MODEL_CLASSIFICATION} | behavioral+safety: {JUDGE_MODEL_BEHAVIORAL}"
    )
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    for sheet, rate in sheet_pass_rates.items():
        tokens = sheet_tokens.get(sheet, 0)
        cost   = sheet_costs.get(sheet, 0.0)
        _ws_append(ws, [run_id, date_str, tag, desc, sheet,
                        round(rate * 100, 1), tokens, round(cost, 4), judge_model, ""])

    overall_scores = list(sheet_pass_rates.values())
    overall_pass   = sum(overall_scores) / len(overall_scores) if overall_scores else 0.0
    overall_tokens = sum(sheet_tokens.values())
    overall_cost   = sum(sheet_costs.values())
    _ws_append(ws, [run_id, date_str, tag, desc, "OVERALL",
                    round(overall_pass * 100, 1), overall_tokens,
                    round(overall_cost, 4), judge_model, ""])
    last_row = _true_last_row(ws)
    for c in range(1, 11):
        ws.cell(last_row, c).font = FONT_BOLD


def _write_run_history_row(ws, run_id: int, tag: str, desc: str,
                            sheet_name: str, pass_rate: float,
                            cost: float, tokens: int, calibrate: bool):
    """Append a single row to Run History — used for both per-sheet and OVERALL entries."""
    judge_model = (
        JUDGE_MODEL_CALIBRATION if calibrate
        else f"classification: {JUDGE_MODEL_CLASSIFICATION} | behavioral+safety: {JUDGE_MODEL_BEHAVIORAL}"
    )
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    _ws_append(ws, [run_id, date_str, tag, desc, sheet_name,
                    round(pass_rate * 100, 1), tokens, round(cost, 4), judge_model, ""])
    if sheet_name == "OVERALL":
        last_row = _true_last_row(ws)
        for c in range(1, 11):
            ws.cell(last_row, c).font = FONT_BOLD


# ---------------------------------------------------------------------------
# Agent response sidecar helpers
# ---------------------------------------------------------------------------

def _sidecar_path(tag: str) -> str:
    return os.path.join(EVAL_RUNS_DIR, tag, "responses.json")


def _save_sidecar(tag: str, responses: dict):
    """Write (or merge-update) the agent response sidecar for a run tag."""
    path = _sidecar_path(tag)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    # Merge with any existing content so incremental per-sheet saves accumulate correctly.
    existing = {}
    if os.path.exists(path):
        try:
            with open(path) as f:
                existing = json.load(f)
        except Exception:
            pass
    existing.update(responses)
    with open(path, "w") as f:
        json.dump(existing, f, indent=2)


def _load_sidecar(tag: str) -> dict:
    """
    Load the agent response sidecar for a run tag.
    Raises SystemExit with a clear message if the file does not exist.
    """
    path = _sidecar_path(tag)
    if not os.path.exists(path):
        sys.exit(
            f"No sidecar found for tag '{tag}' at {path}\n"
            f"Run a full eval with --tag '{tag}' first to generate it."
        )
    with open(path) as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Judge-only helpers
# ---------------------------------------------------------------------------

def _find_latest_tag(wb, exclude_tag: str) -> "str | None":
    """Return the most recent version_tag in Run History, excluding exclude_tag."""
    if RUN_HISTORY_SHEET not in wb.sheetnames:
        return None
    rh = wb[RUN_HISTORY_SHEET]
    vt_col = next(
        (c for c in range(1, rh.max_column + 1) if rh.cell(1, c).value == "version_tag"),
        None,
    )
    if vt_col is None:
        return None
    for row in range(rh.max_row, 1, -1):
        tag = rh.cell(row, vt_col).value
        if tag and tag != "OVERALL" and tag != exclude_tag:
            return str(tag)
    return None


def _find_sheet_tag_cols(ws, tag: str) -> dict:
    """
    Find column indices for a tag's result columns in a test sheet.
    Returns dict with keys: verdict, response, reasoning, failure_reason,
    and optionally escalation_summary. Returns {} if tag columns not found.
    """
    verdict_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v and str(v).startswith(f"{tag} ($"):
            verdict_col = c
            break
    if verdict_col is None:
        return {}
    cols = {
        "verdict":        verdict_col,
        "response":       verdict_col + 1,
        "reasoning":      verdict_col + 2,
        "failure_reason": verdict_col + 3,
    }
    # Check for extra columns (e.g. escalation_summary on Escalation sheet)
    for c in range(verdict_col + 4, ws.max_column + 1):
        hdr = ws.cell(2, c).value
        if hdr and str(hdr).startswith(f"{tag} "):
            suffix = str(hdr)[len(tag) + 1:]
            cols[suffix] = c
    return cols


def _generate_disagreement_report(wb, from_tag: str, new_tag: str, output_path: str):
    """Compare verdicts between from_tag and new_tag; write report to output_path."""
    lines = [
        f"Disagreement report: {from_tag} (baseline) vs {new_tag} (calibration)",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
    ]
    total_disagreements = 0
    sheet_counts = []

    for sheet_name in SHEET_NAMES:
        if sheet_name in PROGRAMMATIC_SHEETS or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        from_cols = _find_sheet_tag_cols(ws, from_tag)
        new_cols  = _find_sheet_tag_cols(ws, new_tag)
        if not from_cols or not new_cols:
            continue

        sheet_lines = []
        for row_idx in range(3, ws.max_row + 1):
            test_id = ws.cell(row_idx, 1).value
            if not test_id:
                continue
            from_v = str(ws.cell(row_idx, from_cols["verdict"]).value or "").upper()
            new_v  = str(ws.cell(row_idx, new_cols["verdict"]).value or "").upper()
            if from_v in ("SKIP", "") or new_v in ("SKIP", ""):
                continue
            if from_v != new_v:
                reasoning = str(ws.cell(row_idx, new_cols["reasoning"]).value or "")[:120]
                sheet_lines.append(
                    f"  {test_id}: baseline={from_v}, {new_tag}={new_v}"
                    f' — "{reasoning}"'
                )
                total_disagreements += 1

        if sheet_lines:
            sheet_counts.append(f"{sheet_name}: {len(sheet_lines)}")
            lines.append(f"=== {sheet_name} ({len(sheet_lines)} disagreements) ===")
            lines.extend(sheet_lines)
            lines.append("")

    if total_disagreements == 0:
        lines.append("No disagreements found — judges agree on all cases.")
    else:
        summary = f"Total disagreements: {total_disagreements}  ({', '.join(sheet_counts)})"
        lines.insert(3, summary)
        lines.insert(4, "")

    with open(output_path, "w") as f:
        f.write("\n".join(lines) + "\n")

    print(f"\nDisagreement report written to: {output_path}")
    print(f"Total disagreements: {total_disagreements}")


async def _run_judge_only(wb, tag: str, desc: str, target_sheets: list,
                           calibrate: bool, yes: bool, delay: float, from_tag: str):
    """Re-judge stored agent responses from a previous run without calling the agent."""
    if not from_tag:
        from_tag = _find_latest_tag(wb, tag)
        if not from_tag:
            sys.exit("--judge-only requires --from-tag or at least one prior run in Run History.")
        print(f"Auto-detected baseline tag: {from_tag}")

    llm_sheets = [s for s in target_sheets if s not in PROGRAMMATIC_SHEETS]
    if not llm_sheets:
        sys.exit("No LLM-judged sheets in target. --judge-only skips Input Guard, Intent Classifier, Output Guard.")

    # Load sidecar — errors clearly if not found (full run required first)
    sidecar = _load_sidecar(from_tag)

    total_cases = 0
    for s in llm_sheets:
        if s in wb.sheetnames:
            ws = wb[s]
            if _find_sheet_tag_cols(ws, from_tag):
                total_cases += ws.max_row - 2

    judge_label = "Opus" if calibrate else "Sonnet"
    print(f"\nJudge-only mode: re-judging up to {total_cases} cases from '{from_tag}' using {judge_label}")
    print(f"Sidecar: {_sidecar_path(from_tag)}  ({len(sidecar)} entries)")
    print(f"Sheets: {', '.join(llm_sheets)}\n")

    if not yes:
        answer = input("Proceed? [y/N] ").strip().lower()
        if answer not in ("y", "yes"):
            print("Aborted.")
            return

    rh_ws  = _ensure_run_history_sheet(wb)
    run_id = _true_last_row(rh_ws)

    sheet_pass_rates: dict = {}
    sheet_costs: dict      = {}
    sheet_tokens: dict     = {}

    for sheet_name in llm_sheets:
        if sheet_name not in wb.sheetnames or sheet_name not in _JUDGE_DISPATCH:
            print(f"[SKIP] {sheet_name} — sheet not found or no judge defined.")
            continue

        ws = wb[sheet_name]
        tag_cols = _find_sheet_tag_cols(ws, from_tag)
        if not tag_cols:
            print(f"[SKIP] {sheet_name} — no columns found for tag '{from_tag}'")
            continue

        print(f"[{sheet_name}] Re-judging stored responses from '{from_tag}'...")
        case_results     = []
        scores           = []
        sheet_judge_cost = 0.0
        judge_fn         = _JUDGE_DISPATCH[sheet_name]

        for row_idx in range(3, ws.max_row + 1):
            test_case = _row_to_dict(ws, row_idx)
            test_id   = test_case.get("test_id", f"row-{row_idx}")

            stored_verdict = str(ws.cell(row_idx, tag_cols["verdict"]).value or "").upper()
            stored_fr      = str(ws.cell(row_idx, tag_cols["failure_reason"]).value or "").lower()
            if stored_verdict == "SKIP" or stored_fr == "api_timeout":
                print(f"  [{test_id}] SKIP (baseline: {stored_verdict or stored_fr})")
                case_results.append({"test_id": test_id, "skipped": True})
                continue

            entry = sidecar.get(test_id, {})
            agent_resp = {
                "response":            entry.get("response", str(ws.cell(row_idx, tag_cols["response"]).value or "")),
                "actions_taken":       entry.get("actions_taken", []),
                "requires_escalation": entry.get("requires_escalation", False),
                "escalation_reason":   entry.get("escalation_reason", ""),
                "confidence":          entry.get("confidence", 0.0),
                "inferred_intent":     entry.get("inferred_intent", ""),
                "context_summary":     entry.get("context_summary", ""),
            }

            if sheet_name == "KB Retrieval":
                ref_titles  = _parse_json_field(test_case.get("reference_articles"), default=[])
                ref_content = await _fetch_kb_reference_content(ref_titles)
                test_case   = {**test_case, "reference_content": ref_content}

            t0 = time.time()
            judgment = await judge_fn(test_case, agent_resp, calibrate)
            latency  = time.time() - t0

            sheet_judge_cost += judgment.get("cost_usd", 0.0)
            verdict = judgment.get("verdict", "fail")
            score   = judgment.get("score", 0.0)
            scores.append(score)

            verdict_label = {"pass": "PASS", "fail": "FAIL"}.get(verdict, "FAIL")
            print(f"  {test_id}: {verdict_label}  ({latency:.1f}s)  ${judgment.get('cost_usd', 0.0):.4f}  {judgment.get('reasoning', '')[:70]}")

            case_results.append({
                "test_id":        test_id,
                "result":         judgment,
                "agent_response": agent_resp,
                "latency_s":      latency,
                "cost_usd":       judgment.get("cost_usd", 0.0),
            })

            if delay > 0:
                time.sleep(delay)

        pass_rate = sum(scores) / len(scores) if scores else 0.0
        sheet_pass_rates[sheet_name] = pass_rate
        sheet_costs[sheet_name]      = sheet_judge_cost
        sheet_tokens[sheet_name]     = 0

        pass_count  = sum(1 for r in case_results if not r.get("skipped") and r.get("result", {}).get("verdict") == "pass")
        total_count = sum(1 for r in case_results if not r.get("skipped"))
        print(f"  → Pass rate: {pass_rate * 100:.1f}%  |  judge: ${sheet_judge_cost:.4f}\n")

        _append_run_column(ws, tag, case_results, sheet_judge_cost,
                           extra_cols=_SHEET_EXTRA_COLS.get(sheet_name))
        _write_run_history_row(rh_ws, run_id, tag, desc, sheet_name,
                               pass_rate, sheet_judge_cost, 0, calibrate)
        wb.save(TEST_CASES_FILE)
        print(f"[{sheet_name}] Complete: {pass_count}/{total_count} passed ({pass_rate * 100:.1f}%) — saved to Run History")

    overall_scores = list(sheet_pass_rates.values())
    overall_pass   = sum(overall_scores) / len(overall_scores) if overall_scores else 0.0
    overall_cost   = sum(sheet_costs.values())
    _write_run_history_row(rh_ws, run_id, tag, desc, "OVERALL",
                           overall_pass, overall_cost, 0, calibrate)

    _build_analysis_sheet(wb)
    for sname in SHEET_NAMES:
        if sname in wb.sheetnames:
            _format_test_sheet(wb[sname])
    if RUN_HISTORY_SHEET in wb.sheetnames:
        _format_run_history_sheet(wb[RUN_HISTORY_SHEET])

    wb.save(TEST_CASES_FILE)
    print(f"Updated {TEST_CASES_FILE} with results.")

    print("\n=== Judge-Only Run Summary ===")
    print(f"Tag:      {tag}")
    print(f"Baseline: {from_tag}")
    if desc:
        print(f"Desc:     {desc}")
    print()
    print(f"  {'Sheet':<25}  {'Pass%':>6}  {'Cost':>8}")
    print(f"  {'-'*25}  {'-'*6}  {'-'*8}")
    for sheet, rate in sheet_pass_rates.items():
        print(f"  {sheet:<25}  {rate * 100:>5.1f}%  ${sheet_costs.get(sheet, 0):.4f}")
    print(f"  {'-'*25}  {'-'*6}  {'-'*8}")
    print(f"  {'OVERALL':<25}  {overall_pass * 100:>5.1f}%  ${overall_cost:.4f}")

    report_path = f"{tag}_disagreements.txt"
    _generate_disagreement_report(wb, from_tag, tag, report_path)


# ---------------------------------------------------------------------------
# Pre-run cost estimation
# ---------------------------------------------------------------------------

def _estimate_cost_per_call(model: str, prompt_tokens: int, completion_tokens: int) -> float:
    """Estimate cost for a call using MODEL_PRICE_PER_TOKEN table."""
    prices = MODEL_PRICE_PER_TOKEN.get(model)
    if not prices:
        return 0.0
    return prompt_tokens * prices["input"] + completion_tokens * prices["output"]


def _estimate_run_cost(wb, target_sheets: list, calibrate: bool) -> tuple:
    """
    Estimate total cost before running.
    Returns (total_cost, breakdown) where breakdown is a list of
    (sheet_name, n_cases, agent_cost, judge_cost, sheet_total) tuples.
    """
    breakdown = []
    total = 0.0

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames or sheet_name not in SHEET_CALL_PROFILE:
            continue
        _ws = wb[sheet_name]
        _hdrs = [_ws.cell(2, c).value for c in range(1, _ws.max_column + 1)]
        _skip_col = next((i + 1 for i, h in enumerate(_hdrs) if h == "skip"), None)
        n_cases = 0
        for _r in range(3, _ws.max_row + 1):
            if not _ws.cell(_r, 1).value:
                break
            if _skip_col:
                _val = str(_ws.cell(_r, _skip_col).value or "").strip().upper()
                if _val in ("TRUE", "1", "YES"):
                    continue
            n_cases += 1
        profile = SHEET_CALL_PROFILE[sheet_name]

        agent_cost = 0.0
        if profile["agent"]:
            per_call = _estimate_cost_per_call(
                AGENT_MODEL,
                AVG_AGENT_PROMPT_TOKENS,
                AVG_AGENT_COMPLETION_TOKENS,
            )
            agent_cost = n_cases * per_call

        judge_cost = 0.0
        if profile["judge"] != "none" and profile["judge_rate"] > 0:
            judge_model = (
                JUDGE_MODEL_CALIBRATION if calibrate
                else JUDGE_MODEL_CLASSIFICATION if profile["judge"] == "classification"
                else JUDGE_MODEL_BEHAVIORAL
            )
            per_judge = _estimate_cost_per_call(
                judge_model,
                AVG_JUDGE_PROMPT_TOKENS,
                AVG_JUDGE_COMPLETION_TOKENS,
            )
            judge_cost = n_cases * profile["judge_rate"] * per_judge

        sheet_total = agent_cost + judge_cost
        total += sheet_total
        breakdown.append((sheet_name, n_cases, agent_cost, judge_cost, sheet_total))

    return total, breakdown


def _print_cost_estimate(breakdown: list, total: float, yes: bool = False) -> bool:
    """Print the pre-run cost estimate and prompt for confirmation. Returns True to proceed."""
    print("\n=== Pre-run cost estimate ===")
    print(f"  {'Sheet':<25} {'Cases':>5}  {'Agent':>8}  {'Judge':>8}  {'Total':>8}")
    print(f"  {'-'*25} {'-'*5}  {'-'*8}  {'-'*8}  {'-'*8}")
    for sheet_name, n_cases, agent_cost, judge_cost, sheet_total in breakdown:
        print(f"  {sheet_name:<25} {n_cases:>5}  ${agent_cost:>7.3f}  ${judge_cost:>7.3f}  ${sheet_total:>7.3f}")
    print(f"  {'-'*25} {'-'*5}  {'-'*8}  {'-'*8}  {'-'*8}")
    print(f"  {'TOTAL':<25} {'':>5}  {'':>8}  {'':>8}  ${total:>7.3f}")
    print()
    print("Note: agent costs are estimates (chars÷4 heuristic); judge costs use LiteLLM pricing.")
    if yes:
        print("\nProceeding automatically (--yes flag set).")
        return True
    answer = input("\nProceed? [y/N] ").strip().lower()
    return answer in ("y", "yes")


# ---------------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------------

async def run_evals(tag: str, desc: str, sheets_filter: list, calibrate: bool,
                    yes: bool = False, delay: float = 5.0,
                    judge_only: bool = False, from_tag: str = ""):
    # 1. Load workbook
    if not os.path.exists(TEST_CASES_FILE):
        sys.exit(f"Test cases file not found: {TEST_CASES_FILE}")

    wb = openpyxl.load_workbook(TEST_CASES_FILE)

    # Determine which sheets to run
    target_sheets = sheets_filter if sheets_filter else SHEET_NAMES
    missing = [s for s in target_sheets if s not in wb.sheetnames]
    if missing:
        sys.exit(f"Sheets not found in workbook: {missing}")

    # Judge-only mode: skip agent, re-judge stored responses from a prior run
    if judge_only:
        await _run_judge_only(wb, tag, desc, target_sheets, calibrate, yes, delay, from_tag)
        return

    # 2. Verify agent is reachable
    print(f"Checking agent at {AGENT_TEST_ENDPOINT} ...")
    try:
        resp = requests.get(
            AGENT_TEST_ENDPOINT.replace("/api/chat/test", "/health"),
            timeout=5,
        )
        if resp.status_code != 200:
            sys.exit(f"Agent health check failed: {resp.status_code}")
        print("Agent is up.\n")
    except requests.exceptions.ConnectionError:
        sys.exit(
            f"Cannot reach agent at {AGENT_TEST_ENDPOINT.replace('/api/chat/test', '')}\n"
            "Make sure it is running with APP_ENV=test."
        )

    # 3. Pre-run cost estimate + confirmation
    estimated_total, cost_breakdown = _estimate_run_cost(wb, target_sheets, calibrate)
    if not _print_cost_estimate(cost_breakdown, estimated_total, yes=yes):
        print("Aborted.")
        return

    # 4. Run each sheet
    all_results: dict[str, list] = {}
    sheet_pass_rates: dict[str, float] = {}
    sheet_costs: dict[str, float] = {}    # actual costs per sheet
    sheet_tokens: dict[str, int] = {}     # actual tokens per sheet
    run_responses: dict = {}              # sidecar: test_id → full agent state

    # Initialise Run History and capture run_id before the loop so per-sheet
    # rows can be written incrementally as each sheet completes.
    rh_ws  = _ensure_run_history_sheet(wb)
    run_id = _true_last_row(rh_ws)

    for sheet_name in target_sheets:
        if sheet_name not in _SHEET_RUNNERS:
            print(f"[SKIP] {sheet_name} — no runner defined.")
            continue

        runner = _SHEET_RUNNERS[sheet_name]
        ws = wb[sheet_name]
        total_rows = sum(1 for r in range(3, ws.max_row + 1) if ws.cell(r, 1).value)
        print(f"[{sheet_name}] Running {total_rows} cases...")

        case_results = []
        scores = []
        sheet_agent_tokens = 0
        sheet_agent_cost   = 0.0
        sheet_judge_cost   = 0.0

        for row_idx in range(3, ws.max_row + 1):
            test_case = _row_to_dict(ws, row_idx)
            test_id = test_case.get("test_id")
            if not test_id:
                break

            if str(test_case.get("skip", "")).strip().upper() in ("TRUE", "1", "YES"):
                print(f"  [{test_id}] SKIP")
                case_results.append({"test_id": test_id, "skipped": True})
                continue

            t0 = time.time()
            agent_resp, judgment = await runner(test_case, calibrate,
                                                test_id=test_id, version_tag=tag)
            latency = time.time() - t0

            # Normalise api_timeout: runners return failure_reason=None on error;
            # set it explicitly so the Analysis sheet can bucket it correctly.
            if judgment.get("verdict") == "fail" and not judgment.get("failure_reason"):
                err = str(agent_resp.get("error", ""))
                if err == "api_timeout" or "timeout" in err.lower():
                    judgment = {**judgment, "failure_reason": "api_timeout"}

            # Accumulate agent token estimates
            p_tok = agent_resp.get("prompt_tokens", 0)
            c_tok = agent_resp.get("completion_tokens", 0)
            sheet_agent_tokens += p_tok + c_tok
            if p_tok + c_tok > 0:
                sheet_agent_cost += _estimate_cost_per_call(
                    AGENT_MODEL, p_tok, c_tok
                )

            # Accumulate judge cost (exact, from LiteLLM)
            sheet_judge_cost += judgment.get("cost_usd", 0.0)

            verdict = judgment.get("verdict", "fail")
            score   = judgment.get("score", 0.0)
            scores.append(score)

            case_cost = judgment.get("cost_usd", 0.0) + _estimate_cost_per_call(
                AGENT_MODEL, p_tok, c_tok
            )
            verdict_label = {"pass": "PASS", "partial": "PART", "fail": "FAIL"}.get(verdict, "FAIL")
            print(f"  {test_id}: {verdict_label}  ({latency:.1f}s)  ${case_cost:.4f}  {judgment.get('reasoning', '')[:70]}")

            case_results.append({
                "test_id": test_id,
                "result": judgment,
                "agent_response": agent_resp,
                "latency_s": latency,
                "cost_usd": case_cost,
            })

            # Accumulate full agent state for sidecar (only for non-error responses)
            if "error" not in agent_resp:
                run_responses[test_id] = {
                    "response":            agent_resp.get("response", ""),
                    "actions_taken":       agent_resp.get("actions_taken", []),
                    "requires_escalation": agent_resp.get("requires_escalation", False),
                    "escalation_reason":   agent_resp.get("escalation_reason", ""),
                    "confidence":          agent_resp.get("confidence", 0.0),
                    "inferred_intent":     agent_resp.get("inferred_intent", ""),
                    "context_summary":     agent_resp.get("context_summary", ""),
                }

            if delay > 0:
                time.sleep(delay)

        pass_rate   = sum(scores) / len(scores) if scores else 0.0
        sheet_total_cost = sheet_agent_cost + sheet_judge_cost
        sheet_pass_rates[sheet_name] = pass_rate
        sheet_costs[sheet_name]      = sheet_total_cost
        sheet_tokens[sheet_name]     = sheet_agent_tokens
        all_results[sheet_name]      = case_results
        print(f"  → Pass rate: {pass_rate * 100:.1f}%  |  agent: ${sheet_agent_cost:.4f}  judge: ${sheet_judge_cost:.4f}  total: ${sheet_total_cost:.4f}\n")

        _append_run_column(ws, tag, case_results, sheet_total_cost,
                           extra_cols=_SHEET_EXTRA_COLS.get(sheet_name))

        # Incremental persistence — save this sheet's results immediately so a
        # mid-run crash doesn't lose completed work.
        pass_count  = sum(1 for r in case_results
                          if not r.get("skipped") and r.get("result", {}).get("verdict") == "pass")
        total_count = sum(1 for r in case_results if not r.get("skipped"))
        _write_run_history_row(rh_ws, run_id, tag, desc, sheet_name,
                               pass_rate, sheet_total_cost, sheet_agent_tokens, calibrate)
        wb.save(TEST_CASES_FILE)
        _save_sidecar(tag, run_responses)
        print(f"[{sheet_name}] Complete: {pass_count}/{total_count} passed ({pass_rate * 100:.1f}%) — saved to Run History")

    # 5. Write OVERALL row to Run History
    overall_scores = list(sheet_pass_rates.values())
    overall_pass   = sum(overall_scores) / len(overall_scores) if overall_scores else 0.0
    overall_tokens = sum(sheet_tokens.values())
    overall_cost   = sum(sheet_costs.values())
    _write_run_history_row(rh_ws, run_id, tag, desc, "OVERALL",
                           overall_pass, overall_cost, overall_tokens, calibrate)

    # 6. Rebuild Analysis sheet, apply formatting, then save
    _build_analysis_sheet(wb)
    for sheet_name in SHEET_NAMES:
        if sheet_name in wb.sheetnames:
            _format_test_sheet(wb[sheet_name])
    if RUN_HISTORY_SHEET in wb.sheetnames:
        _format_run_history_sheet(wb[RUN_HISTORY_SHEET])

    wb.save(TEST_CASES_FILE)
    print(f"Updated {TEST_CASES_FILE} with results.")

    # 7. Print summary
    print("\n=== Run Summary ===")
    print(f"Tag:  {tag}")
    if desc:
        print(f"Desc: {desc}")
    print()
    overall_scores = list(sheet_pass_rates.values())
    overall = sum(overall_scores) / len(overall_scores) if overall_scores else 0.0
    total_actual_cost = sum(sheet_costs.values())
    print(f"  {'Sheet':<25}  {'Pass%':>6}  {'Cost':>8}")
    print(f"  {'-'*25}  {'-'*6}  {'-'*8}")
    for sheet, rate in sheet_pass_rates.items():
        bar = "█" * int(rate * 20) + "░" * (20 - int(rate * 20))
        print(f"  {sheet:<25}  {rate * 100:>5.1f}%  ${sheet_costs.get(sheet, 0):.4f}")
    print(f"  {'-'*25}  {'-'*6}  {'-'*8}")
    print(f"  {'OVERALL':<25}  {overall * 100:>5.1f}%  ${total_actual_cost:.4f}")
    print(f"\n  Estimated: ${estimated_total:.4f}  |  Actual: ${total_actual_cost:.4f}")


def main():
    parser = argparse.ArgumentParser(
        description="Run evals against the AI support agent."
    )
    parser.add_argument(
        "--tag", required=True,
        help="Version tag for this run (e.g. v1.0, v1.1_intent_fix). Used as column header.",
    )
    parser.add_argument(
        "--desc", default="",
        help="Short description of what changed in this run.",
    )
    parser.add_argument(
        "--sheets", default="",
        help="Comma-separated list of sheets to run (default: all). "
             "E.g. 'Input Guard,Intent Classifier'",
    )
    parser.add_argument(
        "--calibrate", action="store_true",
        help="Use the calibration model (Opus) for all judgments instead of the standard tiered approach.",
    )
    parser.add_argument(
        "--yes", "-y", action="store_true",
        help="Skip the cost confirmation prompt and proceed automatically.",
    )
    parser.add_argument(
        "--delay", type=float, default=5.0, metavar="SECONDS",
        help="Seconds to sleep between test cases (default: 5). "
             "Tier 1 Anthropic limit is 50 RPM; each case makes ~4 calls, "
             "so 5s keeps throughput well under the limit.",
    )
    parser.add_argument(
        "--judge-only", action="store_true",
        help="Skip the agent; re-judge stored responses from a previous run. "
             "Skips Input Guard, Intent Classifier, Output Guard (programmatic judges). "
             "Use with --from-tag to specify the baseline run.",
    )
    parser.add_argument(
        "--from-tag", default="",
        help="Baseline tag to read stored agent responses from (used with --judge-only). "
             "Defaults to the most recent tag in Run History.",
    )
    parser.add_argument(
        "--judge-model", default="", metavar="MODEL",
        help="Override the judge model. 'opus' uses claude-opus-4-6 for all judge calls "
             "(equivalent to --calibrate).",
    )
    args = parser.parse_args()

    sheets_filter = [s.strip() for s in args.sheets.split(",") if s.strip()] if args.sheets else []
    calibrate = args.calibrate or (args.judge_model.lower() == "opus")

    asyncio.run(run_evals(
        args.tag, args.desc, sheets_filter, calibrate,
        yes=args.yes, delay=args.delay,
        judge_only=args.judge_only, from_tag=args.from_tag,
    ))


if __name__ == "__main__":
    main()
