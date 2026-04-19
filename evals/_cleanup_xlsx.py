"""
One-time cleanup: trim bloat from eval_test_cases.xlsx and archive old runs.
Run with: python evals/_cleanup_xlsx.py
"""
import re
import os
import sys
import shutil
import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH  = "evals/eval_test_cases.xlsx"
ARCHIVE_PATH = "evals/archived_runs.xlsx"
MAX_RUNS_KEPT = 5

# Sheets that hold test cases with run columns
TEST_CASE_SHEETS = {
    "Input Guard", "Intent Classifier", "Output Guard", "KB Retrieval",
    "Action Execution", "Escalation", "Conversation Quality",
    "PII & Data Leakage", "Policy Compliance", "Graceful Failure",
    "Context Retention",
}
# Sheets to only row/col trim (no run archival)
TRIM_ONLY_SHEETS = {"Analysis", "Run History"}


def find_run_groups(ws):
    """
    Return list of dicts: {start, end, tag} where start/end are 1-based col indices.
    A run group starts at any col where row-2 value matches '<tag> ($N.NNN)'.
    """
    groups = []
    last_col = _true_last_col(ws)
    starts = []
    for c in range(1, last_col + 1):
        v = ws.cell(2, c).value
        if v and re.search(r'\(\$[\d.]+\)', str(v)):
            tag = re.sub(r'\s*\(\$[\d.]+\)\s*$', '', str(v)).strip()
            starts.append((c, tag))

    for i, (start, tag) in enumerate(starts):
        end = starts[i + 1][0] - 1 if i + 1 < len(starts) else last_col
        groups.append({"start": start, "end": end, "tag": tag})
    return groups


def _true_last_row(ws):
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row, 1).value is not None:
            return row
    return 2


def _true_last_col(ws):
    for col in range(ws.max_column, 0, -1):
        if ws.cell(1, col).value is not None or ws.cell(2, col).value is not None:
            return col
    return 1


def unmerge_cols(ws, col_start, col_end):
    """Remove any merged regions that overlap with the given column range."""
    to_remove = []
    for mc in ws.merged_cells.ranges:
        if mc.min_col <= col_end and mc.max_col >= col_start:
            to_remove.append(str(mc))
    for ref in to_remove:
        ws.unmerge_cells(ref)


def copy_group_to_archive(ws_src, group, ws_arc, last_data_row, first_group):
    """Append a run group to the archive sheet, including test_id col if first group."""
    # Find next free column in archive (based on row-2 headers)
    arc_last = 1
    for c in range(ws_arc.max_column, 0, -1):
        if ws_arc.cell(2, c).value is not None or ws_arc.cell(1, c).value is not None:
            arc_last = c
            break

    if first_group and ws_arc.cell(2, 1).value is None:
        # Write test_id column
        for r in range(1, last_data_row + 1):
            ws_arc.cell(r, 1).value = ws_src.cell(r, 1).value
        next_col = 2
    else:
        next_col = arc_last + 1

    width = group["end"] - group["start"] + 1
    for r in range(1, last_data_row + 1):
        for i in range(width):
            src_val = ws_src.cell(r, group["start"] + i).value
            ws_arc.cell(r, next_col + i).value = src_val


def trim_sheet(ws):
    """Delete rows and columns beyond actual data extent."""
    last_row = _true_last_row(ws)
    last_col = _true_last_col(ws)

    extra_rows = ws.max_row - last_row
    if extra_rows > 0:
        ws.delete_rows(last_row + 1, extra_rows)

    extra_cols = ws.max_column - last_col
    if extra_cols > 0:
        unmerge_cols(ws, last_col + 1, ws.max_column)
        ws.delete_cols(last_col + 1, extra_cols)

    # Remove column style definitions beyond true extent
    from openpyxl.utils import column_index_from_string
    stale_col_keys = [
        k for k in list(ws.column_dimensions.keys())
        if column_index_from_string(k) > last_col
    ]
    for k in stale_col_keys:
        del ws.column_dimensions[k]

    # Purge styled-but-empty ghost cells beyond true extent from ws._cells.
    # delete_cols removes cell values but leaves cell objects with styling, which
    # inflate calculate_dimension() and cause Excel to allocate memory for them.
    stale_keys = [
        k for k in list(ws._cells.keys())
        if k[0] > last_row or k[1] > last_col
    ]
    for k in stale_keys:
        del ws._cells[k]


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)

    # Load or create archive workbook
    if os.path.exists(ARCHIVE_PATH):
        wb_arc = openpyxl.load_workbook(ARCHIVE_PATH)
    else:
        wb_arc = openpyxl.Workbook()
        wb_arc.remove(wb_arc.active)  # remove default empty sheet

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[{sheet_name}] before: rows={ws.max_row}, cols={ws.max_column}")

        if sheet_name in TRIM_ONLY_SHEETS:
            trim_sheet(ws)
            print(f"  → trimmed to rows={ws.max_row}, cols={ws.max_column}")
            continue

        if sheet_name not in TEST_CASE_SHEETS:
            continue

        last_data_row = _true_last_row(ws)
        groups = find_run_groups(ws)
        print(f"  {len(groups)} run groups: {[g['tag'] for g in groups]}")

        to_archive = groups[:-MAX_RUNS_KEPT] if len(groups) > MAX_RUNS_KEPT else []
        to_keep    = groups[-MAX_RUNS_KEPT:] if len(groups) > MAX_RUNS_KEPT else groups

        if to_archive:
            if sheet_name not in wb_arc.sheetnames:
                ws_arc = wb_arc.create_sheet(sheet_name)
            else:
                ws_arc = wb_arc[sheet_name]

            for idx, group in enumerate(to_archive):
                copy_group_to_archive(ws, group, ws_arc, last_data_row, first_group=(idx == 0))
            print(f"  → archived {len(to_archive)} run(s): {[g['tag'] for g in to_archive]}")

            # Delete archived columns from source, right to left to preserve indices
            for group in reversed(to_archive):
                width = group["end"] - group["start"] + 1
                unmerge_cols(ws, group["start"], group["end"])
                ws.delete_cols(group["start"], width)
                print(f"  → deleted cols {group['start']}-{group['end']} ({group['tag']})")

        # Now trim rows/cols beyond actual data (after column deletions)
        trim_sheet(ws)

        print(f"  → after: rows={ws.max_row}, cols={ws.max_column}")
        kept_tags = [g["tag"] for g in (to_keep if to_archive else groups)]
        print(f"  → kept runs: {kept_tags}")

    # Re-apply column visibility after column shifts (cleanup changes column positions)
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from evals.run_evals import _apply_column_visibility
    for sheet_name in TEST_CASE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        latest_tag = None
        for c in range(ws.max_column, 0, -1):
            h = ws.cell(2, c).value
            if h and ' ($' in str(h):
                latest_tag = str(h).rsplit(' ($', 1)[0]
                break
        if latest_tag:
            _apply_column_visibility(ws, latest_tag)

    wb.save(XLSX_PATH)
    if wb_arc.sheetnames:
        wb_arc.save(ARCHIVE_PATH)
        print(f"\nArchived runs saved to {ARCHIVE_PATH}")
    print(f"\nSaved cleaned workbook to {XLSX_PATH}")

    # Verify final dimensions
    wb2 = openpyxl.load_workbook(XLSX_PATH)
    print("\n=== Final dimensions ===")
    for name in wb2.sheetnames:
        ws2 = wb2[name]
        print(f"  {name}: rows={ws2.max_row}, cols={ws2.max_column}")


if __name__ == "__main__":
    main()
